import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook

# CONFIG
TEMPLATE_PATH = "TOOL_template.xlsm"
SHEET_NAME = "Trackingnummern"
START_ROW = 7
COLUMN = "A"

st.set_page_config(page_title="TOOL ELEGGIBILI", page_icon="📦")
st.title("TOOL ELEGGIBILI – Weekend Eligible")

st.write(
    "Carica uno o più file ULD Detail (.xlsx). "
    "Il tool estrarrà i tracking con WeekendEligibleVolume = 1 "
    "e li inserirà nel template TOOL (.xlsm) mantenendo le macro."
)

uploaded_files = st.file_uploader(
    "Carica i file ULD Detail",
    type=["xlsx"],
    accept_multiple_files=True
)


if uploaded_files:
    if st.button("Elabora e genera TOOL ELEGGIBILI"):
        all_tracks = []
        all_colH_values = []   # <-- NUOVO: valori colonna H

        # 1) Estrazione tracking + colonna H da tutti i file
        for file in uploaded_files:
            df = pd.read_excel(file, header=5)  # riga 6 come intestazione

            # Controllo colonne tracking
            missing = []
            if "WeekendEligibleVolume" not in df.columns:
                missing.append("WeekendEligibleVolume")
            if "TrackingNumber" not in df.columns:
                missing.append("TrackingNumber")

            if missing:
                st.warning(
                    f"File {file.name} ignorato: colonne mancanti {', '.join(missing)}."
                )
                continue

            # --- Tracking eleggibili ---
            df_filtered = df[df["WeekendEligibleVolume"] == 1]
            tracks = df_filtered["TrackingNumber"].dropna().astype(str).tolist()
            all_tracks.extend(tracks)

            # --- Colonna H (valori generici) ---
            if len(df.columns) >= 8:
                col_H = df.iloc[6:, 7]  # colonna H dalla riga 7
                for val in col_H:
                    if pd.notna(val):
                        all_colH_values.append(str(val))

        # Rimuovi duplicati mantenendo l'ordine
        seen = set()
        unique_tracks = []
        for t in all_tracks:
            if t not in seen:
                seen.add(t)
                unique_tracks.append(t)

        if not unique_tracks:
            st.warning("Nessun tracking con WeekendEligibleVolume = 1 trovato nei file caricati.")
            st.stop()

        st.success(f"Trovati {len(unique_tracks)} tracking eleggibili.")
        st.success(f"Trovati {len(all_colH_values)} valori nella colonna H.")

        # 2) Carica il TEMPLATE XLSM con macro
        try:
            wb = load_workbook(TEMPLATE_PATH, keep_vba=True)
        except Exception as e:
            st.error(f"Impossibile aprire il template TOOL: {e}")
            st.stop()

        if SHEET_NAME not in wb.sheetnames:
            st.error(f"Nel template non esiste il foglio '{SHEET_NAME}'.")
            st.stop()

        ws = wb[SHEET_NAME]

        # 3) Pulisci colonna A da riga 7 in giù
        max_row = ws.max_row
        for r in range(START_ROW, max_row + 1):
            ws[f"{COLUMN}{r}"] = None

        # 4) Scrivi i tracking
        row = START_ROW
        for t in unique_tracks:
            ws[f"{COLUMN}{row}"] = t
            row += 1

        # 5) Salva direttamente come XLSM mantenendo le macro
        final_buffer = BytesIO()
        wb.save(final_buffer)
        final_buffer.seek(0)

        # 6) Nome file XLSM
        today_str = datetime.now().strftime("%d.%m.%Y")
        filename_xlsm = f"TOOL ELEGGIBILI {today_str}.xlsm"

        # 7) Download XLSM
        st.download_button(
            "Scarica TOOL ELEGGIBILI",
            data=final_buffer.getvalue(),
            file_name=filename_xlsm,
            mime="application/vnd.ms-excel.sheet.macroEnabled.12"
        )

        # ---------------------------------------------------------
        # 8) GENERAZIONE FILE TXT CON I VALORI DELLA COLONNA H
        # ---------------------------------------------------------

        txt_buffer = BytesIO()
        txt_content = "\n".join(all_colH_values)
        txt_buffer.write(txt_content.encode("utf-8"))
        txt_buffer.seek(0)

        filename_txt = f"COLONNA_H_{today_str}.txt"

        st.download_button(
            "Scarica TXT – Colonna H",
            data=txt_buffer.getvalue(),
            file_name=filename_txt,
            mime="text/plain"
        )
